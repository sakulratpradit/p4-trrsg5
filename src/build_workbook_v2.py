#!/usr/bin/env python3
"""Build Pillar4_US_Portfolio_Workbook.xlsx from portfolio_data.py.

v3 rebuild (static values, no external formulas — the original v2 builder with
STOCKHISTORY live formulas was lost in a container reset; this version derives
everything from portfolio_data.py so it always ties to the dashboard exactly).

Sheets: Summary | Holdings | Fundamentals | Trades | Monthly | Sold & Realized
Usage:  python3 build_workbook_v2.py
"""
import portfolio_data as P
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HDR = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="1F3864")
NUM2 = '#,##0.00'
NUM0 = '#,##0'
PCT1 = '+0.0%;-0.0%'

def sheet(wb, title, headers, widths):
    ws = wb.create_sheet(title)
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font, cell.fill = HDR, FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    return ws

def main():
    wb = Workbook(); wb.remove(wb.active)
    mv = lambda s: (P.POS[s['t']]['shares'] or 0) * (s['price'] or 0) if P.POS[s['t']].get('shares') is not None and s.get('price') is not None else None

    # ---- Holdings ----
    ws = sheet(wb, "Holdings",
               ["Ticker","Name","Group","Shares","Cost (USD)","Avg Cost","Price","Mkt Value","Unreal P/L","Unreal %","FX (THB)","Cost (THB)","MV (THB)"],
               [9,24,22,9,13,11,11,13,12,10,10,14,14])
    r = 2
    for s in P.STOCKS:
        p = P.POS[s['t']]
        if p.get('shares') is None: continue
        m = mv(s); cost = p.get('cost'); fx = P.THBFX.get(s['t'], P.TOTALS.get('fx'))
        un = (m - cost) if (m is not None and cost is not None) else None
        row = [s['t'], s['name'], P.GROUPS[s['g']], p['shares'], cost,
               (cost/p['shares'] if cost and p['shares'] else None), s.get('price'), m, un,
               (un/cost if un is not None and cost else None), fx,
               (cost*fx if cost and fx else None), (m*P.TOTALS.get('fx') if m and P.TOTALS.get('fx') else None)]
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c in (5,6,7,8,9,11): cell.number_format = NUM2
            if c in (12,13): cell.number_format = NUM0
            if c == 10: cell.number_format = PCT1
        r += 1
    tot_cost = sum(P.POS[s['t']].get('cost') or 0 for s in P.STOCKS)
    tot_mv = sum(mv(s) or 0 for s in P.STOCKS)
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    for c, v, f in ((5, tot_cost, NUM2), (8, tot_mv, NUM2), (9, tot_mv-tot_cost, NUM2), (10, (tot_mv-tot_cost)/tot_cost, PCT1)):
        cell = ws.cell(row=r, column=c, value=v); cell.font = Font(bold=True); cell.number_format = f

    # ---- Fundamentals ----
    cols = [("t","Ticker",9),("name","Name",24),("price","Price",10),("mcapB","MCap $B",10),
            ("pe","P/E",8),("fpe","Fwd P/E",9),("ps","P/S",8),("peg","PEG",8),
            ("gm","Gross M%",9),("pm","Profit M%",9),("revB","Rev $B",9),("revG","Rev G%",8),
            ("eps","EPS",9),("epsG","EPS G%",9),("roi","ROIC%",8),("roe","ROE%",8),
            ("fcfB","FCF $B",9),("capexB","Capex $B",9),("ath","52w High",10),
            ("fvMin","Tgt Min",9),("fvAvg","Tgt Avg",9),("fvMax","Tgt Max",9),("an","Analysts",9),
            ("r40","Rule of 40",10)]
    ws = sheet(wb, "Fundamentals", [c[1] for c in cols] + ["P/FCF","EV/EBITDA","D/E"], [c[2] for c in cols] + [9,10,8])
    for r, s in enumerate(P.STOCKS, 2):
        m3 = P.M3.get(s['t'], {})
        vals = [s.get(k) for k, _, _ in cols] + [m3.get('pfcf'), m3.get('ev'), m3.get('de')]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c > 2 and isinstance(v, (int, float)): cell.number_format = NUM2

    # ---- Trades ----
    ws = sheet(wb, "Trades", ["Date","Action","Ticker","Shares","Price","Amount","Est.","Note"], [13,8,9,9,11,12,6,60])
    for r, t in enumerate(P.TRADES, 2):
        for c, v in enumerate([t.get('date'), t.get('action'), t.get('t'), t.get('shares'),
                               t.get('price'), t.get('amount'), t.get('est'), t.get('note') or t.get('cash') or ""], 1):
            cell = ws.cell(row=r, column=c, value=v if not isinstance(v, dict) else str(v))
            if c in (5,6): cell.number_format = NUM2

    # ---- Monthly ----
    ws = sheet(wb, "Monthly", ["Month","Buy USD","Buy Lots","Sell USD","Sell Lots","Realized USD","Realized THB"], [10,13,9,13,9,13,13])
    for r, m in enumerate(P.MONTHLY, 2):
        for c, v in enumerate([m['ym'], m['buyUSD'], m['buyLots'], m['sellUSD'], m['sellLots'], m['realizedUSD'], m['realizedTHB']], 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c in (2,4,6,7): cell.number_format = NUM2

    # ---- Sold & Realized ----
    ws = sheet(wb, "Sold & Realized", ["Ticker","Qty","Cost USD","Cost THB","Realized USD","Realized THB","Group"], [9,8,12,13,13,13,22])
    for r, s in enumerate(P.SOLD, 2):
        for c, v in enumerate([s['t'], s.get('qty'), s.get('costUSD'), s.get('costTHB'), s.get('glUSD'), s.get('glTHB'), P.GROUPS[s['g']] if s.get('g') is not None and s['g'] < len(P.GROUPS) else ""], 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c in (3,4,5,6): cell.number_format = NUM2

    # ---- Summary (first tab) ----
    ws = wb.create_sheet("Summary", 0)
    ws.column_dimensions['A'].width = 34; ws.column_dimensions['B'].width = 20
    realized = P.TOTALS.get('realizedUSD') or sum(x['amount'] for x in P.REALIZED)
    lines = [("Pillar 4 — US Portfolio", None), ("As of", P.ASOF),
             ("Held positions", sum(1 for s in P.STOCKS if P.POS[s['t']].get('shares') is not None)),
             ("Held cost (USD)", round(tot_cost, 2)), ("Market value (USD)", round(tot_mv, 2)),
             ("Unrealized P/L (USD)", round(tot_mv - tot_cost, 2)),
             ("Unrealized %", f"{(tot_mv-tot_cost)/tot_cost*100:+.2f}%"),
             ("Realized P/L (USD)", round(realized, 2)),
             (f"Cash — {P.CASH['broker']}", P.CASH['amount']),
             ("Cash as of", P.CASH['asof'])]
    for r, (k, v) in enumerate(lines, 1):
        ws.cell(row=r, column=1, value=k).font = Font(bold=(r == 1 or v is None))
        if v is not None:
            cell = ws.cell(row=r, column=2, value=v)
            if isinstance(v, float): cell.number_format = NUM2

    wb.save("Pillar4_US_Portfolio_Workbook.xlsx")
    print("Pillar4_US_Portfolio_Workbook.xlsx written")

if __name__ == "__main__":
    main()
