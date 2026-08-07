#!/usr/bin/env python3
"""Build the Stock Notes workbook (Pillar 4 US portfolio).

Sources
  1. src/stock_notes_archive.json  - frozen chat commentary mined from the
     Jul 20 - Aug 7 2026 session, plus dated dashboard-note segments.
  2. git history of index.html     - every ASOF note ever written, re-read on
     each run so new daily commentary flows in automatically.
  3. src/portfolio_data.py         - current price / group / position / budget.

The "Our Notes" sheet is NEVER overwritten: if the output workbook already
exists its rows are read back and carried forward verbatim.

Usage:  python3 src/gen_stock_notes.py [output.xlsx]
"""
import csv, json, os, re, subprocess, sys, datetime

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
sys.path.insert(0, HERE)
import portfolio_data as P

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = 'Arial'
HDR_FILL = PatternFill('solid', fgColor='1F3864')
HDR_FONT = Font(name=FONT, size=10, bold=True, color='FFFFFF')
SUB_FILL = PatternFill('solid', fgColor='D9E2F3')
YELLOW = PatternFill('solid', fgColor='FFFF00')
THIN = Side(style='thin', color='BFBFBF')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPWRAP = Alignment(vertical='top', wrap_text=True)
TOP = Alignment(vertical='top')

RISKY = {'BE','SE','TE','NU','PL','FI','MA','GM','PE','PM','EPS','AI','US','EV','ATM','SEC','IPO',
         'GAAP','TTM','ROE','ROIC','FY','Q1','Q2','Q3','Q4','OHLC','WACC','NOT','AND','THE','ALL',
         'FOR','BUY','NOTE','ON','IT','NOW','APP','NET','V'}


LINK = re.compile(r'\[([^\]]{1,120})\]\((https?://[^)]+)\)')


def clean(txt):
    """Markdown -> plain text that reads properly in a spreadsheet cell."""
    t = LINK.sub(r'\1 (\2)', txt)
    t = t.replace('**', '').replace('__', '')
    t = re.sub(r'^#+\s*', '', t)
    t = re.sub(r'[ \t]+', ' ', t).strip()
    return t


def is_low_signal(txt):
    """Drop link dumps, bare source lists and table separator rows."""
    if re.match(r'^\s*(\*\*)?Sources?\s*:', txt, re.I):
        return True
    if txt.count('](http') >= 2:
        return True
    if re.match(r'^\s*\|[\s\-:|]+\|\s*$', txt):
        return True
    letters = sum(ch.isalpha() for ch in txt)
    return letters < 40


def git_asof_segments():
    """Every distinct ' | '-delimited ASOF segment across index.html's history."""
    try:
        log = subprocess.check_output(
            ['git', '-C', REPO, 'log', '--format=%h|%cI', '--reverse', '--', 'index.html']
        ).decode().strip().split('\n')
    except Exception:
        return []
    seen, order = {}, []
    for line in log:
        if '|' not in line:
            continue
        h, dt = line.split('|', 1)
        try:
            blob = subprocess.check_output(['git', '-C', REPO, 'show', f'{h}:index.html'],
                                           stderr=subprocess.DEVNULL).decode('utf-8', 'replace')
        except Exception:
            continue
        m = re.search(r'^const ASOF = (".*"|\'.*\');$', blob, re.M)
        if not m:
            continue
        raw = m.group(1)
        try:
            s = json.loads(raw) if raw.startswith('"') else raw[1:-1]
        except Exception:
            s = raw[1:-1]
        for seg in [x.strip() for x in s.split(' | ') if x.strip()]:
            key = seg[:80]
            if key not in seen:
                seen[key] = {'date': dt[:10], 'commit': h, 'text': seg}
                order.append(key)
            elif len(seg) > len(seen[key]['text']):
                seen[key]['text'] = seg
    return [seen[k] for k in order]


def attribute(segments, tickers):
    """Tag dashboard-note segments to tickers.

    A segment that names more than 8 tickers is a bulk daily-price-refresh note:
    it says nothing specific about any one name, so it is left out of the
    per-ticker log (it still appears in full on the Dashboard Notes sheet).
    """
    tset, tok = set(tickers), re.compile(r'\b([A-Z]{2,5})\b')
    out = {t: [] for t in tickers}
    for s in segments:
        hits = {m.group(1) for m in tok.finditer(s['text'])
                if m.group(1) in tset and m.group(1) not in RISKY}
        if len(hits) > 8:
            continue
        if re.match(r'^[A-Z][a-z]{2} \d{1,2}, 20\d\d', s['text']) and len(s['text']) < 120:
            continue
        for t in hits:
            out[t].append({'ts': s['date'], 'text': s['text'],
                           'src': 'Dashboard note ' + s['commit']})
    return out


def read_our_notes_csv():
    """Durable copy of Salee & Dad's own notes, committed in the repo.

    A fresh scheduled run has no previous workbook on disk, so this CSV is what
    makes the Our Notes sheet survive from one daily rebuild to the next.
    Columns: Date, Who, Ticker, Note, Action, Done
    """
    path = os.path.join(HERE, 'our_notes.csv')
    if not os.path.exists(path):
        return []
    rows = []
    with open(path, newline='', encoding='utf-8') as f:
        for i, r in enumerate(csv.reader(f)):
            if i == 0 and r and r[0].strip().lower() == 'date':
                continue
            if any(c.strip() for c in r):
                rows.append((r + [''] * 6)[:6])
    return rows


def read_existing_our_notes(path):
    if not os.path.exists(path):
        return []
    try:
        wb = load_workbook(path)
        if 'Our Notes' not in wb.sheetnames:
            return []
        ws, rows = wb['Our Notes'], []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if any(c not in (None, '') for c in r):
                rows.append([('' if c is None else c) for c in r][:6])
        return rows
    except Exception:
        return []


def style_header(ws, headers, widths, height=28):
    ws.append(headers)
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=i)
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = height
    ws.freeze_panes = 'A2'


def build(out_path):
    archive_path = os.path.join(HERE, 'stock_notes_archive.json')
    archive = json.load(open(archive_path)) if os.path.exists(archive_path) else {'chat': {}}
    chat = archive.get('chat', {})

    stocks = sorted(P.STOCKS, key=lambda s: (s['g'], s['t']))
    tickers = [s['t'] for s in stocks]
    segments = git_asof_segments()
    asof_by = attribute(segments, tickers)

    chat_clean = {}
    for t in tickers:
        kept, seen = [], set()
        for n in sorted(chat.get(t, []), key=lambda x: x['ts']):
            if is_low_signal(n['text']):
                continue
            c = clean(n['text'])
            if c in seen:
                continue
            seen.add(c)
            kept.append({'ts': n['ts'], 'text': c})
        chat_clean[t] = kept

    carried = read_our_notes_csv()
    seen_rows = {tuple(str(c).strip() for c in r) for r in carried}
    for r in read_existing_our_notes(out_path):
        k = tuple(str(c).strip() for c in r)
        if k not in seen_rows:
            seen_rows.add(k)
            carried.append(r)
    today = datetime.date.today().isoformat()

    wb = Workbook()

    # ---------------- Read Me ----------------
    ws = wb.active
    ws.title = 'Read Me'
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 108
    rows = [
        ('PILLAR 4 - US STOCK NOTES', ''),
        ('', ''),
        ('What this is',
         'Every comment Claude has written about each stock in the US dashboard, kept in one place so '
         'you and Dad can re-read the reasoning instead of asking the same question twice.'),
        ('Generated', today),
        ('Objective on record', 'SL-P4: 100% capital gain over five years, 2026-2030 (about 15% a year).'),
        ('', ''),
        ('SHEET', 'WHAT IS IN IT'),
        ('Stock Index',
         'One row per ticker: group, whether you hold it, shares, cost, budget, last stored price, how '
         'many comments exist, and the most recent thing Claude said about it. Start here.'),
        ('Comment Log',
         'The full archive. One row per comment, tagged with the ticker and the date. Use the filter '
         'arrow on the Ticker column to pull up a single stock.'),
        ('Dashboard Notes',
         'The dated technical notes stored inside the dashboard itself - data conflicts, earnings dates, '
         'warnings, what was verified and what was not. Longer and more detailed than the chat comments.'),
        ('Our Notes',
         'YOURS. Claude never overwrites this sheet - it is read back and carried forward on every '
         'rebuild. Put your own observations, Dad\'s notes, and decisions here.'),
        ('Follow-ups', 'Open items Claude is carrying: unanswered offers, data to fix, earnings to watch.'),
        ('', ''),
        ('HOW TO LOOK UP ONE STOCK', ''),
        ('1', 'Go to the Comment Log sheet.'),
        ('2', 'Click the small arrow in the Ticker header (row 1, column A).'),
        ('3', 'Untick "Select All", tick just the ticker you want, press OK.'),
        ('4', 'Every comment about that stock, oldest to newest, is now on screen.'),
        ('5', 'To clear it, click the arrow again and tick "Select All".'),
        ('', ''),
        ('IMPORTANT LIMITS', ''),
        ('Not advice',
         'These are analytical comments, not licensed investment advice. Every order is placed by Salee.'),
        ('Comments age',
         'A comment is a snapshot of what was true and known on its date. A buy case from three weeks '
         'ago may have been overtaken by an earnings report. Always check the date column.'),
        ('Prices are closes',
         'Prices in the Stock Index come from the dashboard and are closing prices with a stated date, '
         'not live quotes.'),
        ('Coverage',
         'Comments were mined from the chat history of the Jul 20 - Aug 7 2026 session plus the full '
         'git history of the dashboard. Anything said before that session is not captured.'),
    ]
    for a, b in rows:
        ws.append([a, b])
    ws['A1'].font = Font(name=FONT, size=16, bold=True, color='1F3864')
    for r in (7, 14, 21):
        for col in ('A', 'B'):
            ws[f'{col}{r}'].fill = SUB_FILL
            ws[f'{col}{r}'].font = Font(name=FONT, size=10, bold=True)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            if c.font.size != 16 and not c.font.bold:
                c.font = Font(name=FONT, size=10)
            c.alignment = TOPWRAP
    for r in range(3, ws.max_row + 1):
        ws.row_dimensions[r].height = None

    # ---------------- Comment Log ----------------
    log = wb.create_sheet('Comment Log')
    style_header(log, ['Ticker', 'Date', 'Source', 'Comment'], [10, 12, 22, 150])
    entries = []
    for t in tickers:
        for n in chat_clean.get(t, []):
            entries.append((t, n['ts'], 'Claude - chat', n['text']))
        for n in asof_by.get(t, []):
            entries.append((t, n['ts'], n['src'], clean(n['text'])))
    entries.sort(key=lambda e: (e[0], e[1]))
    for e in entries:
        log.append(list(e))
    for row in log.iter_rows(min_row=2):
        for c in row:
            c.font = Font(name=FONT, size=10)
            c.alignment = TOPWRAP
            c.border = BOX
    if log.max_row > 1:
        log.auto_filter.ref = f'A1:D{log.max_row}'

    # ---------------- Stock Index ----------------
    idx = wb.create_sheet('Stock Index', 1)
    style_header(idx, ['Ticker', 'Company', 'Group', 'Held', 'Shares', 'Cost (USD)',
                       'Budget (USD)', 'Last close (USD)', 'Price date', 'Comments',
                       'First comment', 'Last comment', 'Most recent comment'],
                 [10, 26, 26, 7, 10, 13, 13, 14, 12, 11, 13, 13, 110])
    for s in stocks:
        t = s['t']
        pos = P.POS.get(t, {})
        cnotes = chat_clean.get(t, [])
        anotes = [{'ts': n['ts'], 'text': clean(n['text'])} for n in asof_by.get(t, [])]
        notes = sorted(cnotes + anotes, key=lambda n: n['ts'])
        # prefer Claude's own written analysis, and prefer a substantial paragraph
        pool = [n for n in cnotes if len(n['text']) > 220] or cnotes or notes
        latest = pool[-1]['text'] if pool else 'No comment on record yet.'
        if len(latest) > 900:
            latest = latest[:900].rsplit(' ', 1)[0] + ' ...'
        r = idx.max_row + 1
        idx.append([
            t, s['name'], P.GROUPS[s['g']],
            'YES' if pos.get('shares') else '',
            pos.get('shares'), pos.get('cost'), pos.get('budget'),
            s.get('price'), s.get('pxd'),
            f"=COUNTIF('Comment Log'!$A:$A,$A{r})",
            notes[0]['ts'] if notes else '', notes[-1]['ts'] if notes else '',
            latest,
        ])
    for row in idx.iter_rows(min_row=2):
        for c in row:
            c.font = Font(name=FONT, size=10)
            c.border = BOX
            c.alignment = TOPWRAP if c.column == 13 else TOP
        row[3].alignment = Alignment(horizontal='center', vertical='top')
        if row[3].value == 'YES':
            row[3].font = Font(name=FONT, size=10, bold=True, color='006100')
        for i in (5,):
            row[i].number_format = '#,##0.00'
        for i in (6, 7):
            row[i].number_format = '$#,##0;($#,##0);-'
        row[4].number_format = '#,##0.##'
    idx.auto_filter.ref = f'A1:M{idx.max_row}'

    # ---------------- Dashboard Notes ----------------
    dn = wb.create_sheet('Dashboard Notes')
    style_header(dn, ['Date', 'Commit', 'Note stored in the dashboard'], [12, 11, 165])
    for s in segments:
        dn.append([s['date'], s['commit'], clean(s['text'])])
    for row in dn.iter_rows(min_row=2):
        for c in row:
            c.font = Font(name=FONT, size=10)
            c.alignment = TOPWRAP
            c.border = BOX
    if dn.max_row > 1:
        dn.auto_filter.ref = f'A1:C{dn.max_row}'

    # ---------------- Our Notes ----------------
    on = wb.create_sheet('Our Notes')
    style_header(on, ['Date', 'Who', 'Ticker', 'Our note / question', 'Action agreed', 'Done?'],
                 [12, 12, 10, 80, 50, 9])
    if carried:
        for r in carried:
            on.append(r + [''] * (6 - len(r)))
    else:
        on.append(['2026-08-07', 'Salee', 'NVDA',
                   'EXAMPLE ROW - delete me. Dad wants to add on any dip under 210; check budget headroom first.',
                   'Wait for 210 or lower, then buy up to budget.', 'No'])
        on.cell(row=2, column=1).fill = YELLOW
    for row in on.iter_rows(min_row=2):
        for c in row:
            c.font = Font(name=FONT, size=10)
            c.alignment = TOPWRAP
            c.border = BOX
    on.auto_filter.ref = f'A1:F{max(on.max_row, 2)}'
    note = on.cell(row=max(on.max_row, 2) + 2, column=1)
    note.value = ('This sheet belongs to Salee and Dad. Claude reads it back and copies it forward '
                  'every time the workbook is rebuilt, so nothing typed here is ever lost.')
    note.font = Font(name=FONT, size=9, italic=True, color='808080')

    # ---------------- Follow-ups ----------------
    fu = wb.create_sheet('Follow-ups')
    style_header(fu, ['Ticker', 'Type', 'Open item', 'Raised'], [10, 20, 130, 12])
    for r in FOLLOW_UPS:
        fu.append(list(r))
    for row in fu.iter_rows(min_row=2):
        for c in row:
            c.font = Font(name=FONT, size=10)
            c.alignment = TOPWRAP
            c.border = BOX
    fu.auto_filter.ref = f'A1:D{fu.max_row}'

    wb.save(out_path)
    return out_path, len(entries), len(tickers), len(segments)


FOLLOW_UPS = [
    ('ZETA', 'Data error', "Stored all-time high 25.95 is BELOW the stored price 27.07 and below the sourced 52-week high 28.50 - impossible. 52-week low should be 14.37 (stored 12.10) and high 28.50 (stored 27.07). Needs a sourced ATH before it can be fixed; Claude will not guess it.", '2026-08-06'),
    ('AIRJ', 'Data check', 'Stored all-time high 29.27 conflicts with a researcher figure of 49.11 (Mar 11 2024, TradingView). TradingView is on the unreliable list, so a second source is needed before changing anything.', '2026-08-05'),
    ('APP', 'Data gap', 'Capital expenditure is stored blank. Fill it if a source turns up.', '2026-08-05'),
    ('SKHY', 'Stale data', 'Fundamentals are pre-Q2 and therefore flattering. Budget left about $8,087.', '2026-08-04'),
    ('LWLG / AXTI', 'Chart scale', "Absurd-scale valuations distort three charts: LWLG price-to-sales 5,040x, AXTI price-to-earnings 957x and EV/EBITDA 1,042x. Claude has offered to cap the display scale so the other bars stay readable - awaiting a yes.", '2026-08-06'),
    ('AXTI', 'Risk to watch', "Beijing Tongmei withdrew its Shanghai STAR Market listing on Jul 8 2026, triggering redemption rights on about US$49M held by eleven private-equity funds. Company says it has the cash.", '2026-08-06'),
    ('COHR', 'Earnings', 'Reports FY26 Q4 and full year on Aug 12, 2026.', '2026-08-05'),
    ('LWLG', 'Earnings', 'Reports Q2 on Aug 11, 2026 (company-confirmed).', '2026-08-05'),
    ('NVDA', 'Earnings', 'Reports Q2 FY2027 on Aug 26, 2026 (company-confirmed). Budget headroom about $2,668, roughly 12 shares.', '2026-08-06'),
    ('AIRJ', 'Earnings', 'Next earnings estimated Aug 21, 2026 - NOT company-confirmed.', '2026-08-05'),
    ('AXTI', 'Earnings', 'Next earnings date could not be found and is not company-confirmed.', '2026-08-06'),
    ('LEU', 'Idea pending', "On Dad's Aug 4 list at or below $150; would sit in Nuclear Energy. Red flags on record: forward P/E 81 above trailing P/E 65, revenue -4.1%, EPS -55.5%, short interest 25.4%.", '2026-08-04'),
    ('AMP', 'Needs clarifying', "Item 6 on Dad's Aug 04 note reads 'AMP <= $350'. Most likely AMD. Awaiting confirmation.", '2026-08-04'),
    ('TER', 'Proposal pending', "Claude proposed moving TER from Robotic/Automation/EV to Chips Manufacturing and renaming the Photonic group to 'Interconnect & Networking'. Awaiting a decision.", '2026-08-01'),
    ('FISV', 'Housekeeping', 'FI has been renamed FISV. The rename still needs to be applied to the dashboard.', '2026-07-30'),
    ('XE / YUM', 'Price check', 'Both were stuck at their Aug 3 close at one point. The Aug 5 refresh may have fixed it - unverified.', '2026-08-05'),
    ('NNE', 'Data check', 'History shows a Jul 29 low of 14.71; stored 52-week low may need updating.', '2026-08-05'),
    ('Portfolio', 'Reconciliation', 'Monthly-flow and totals figures differ by about $5,636 - a pre-existing inconsistency never chased down.', '2026-07-28'),
    ('Portfolio', 'Open offer', 'Claude has offered but not yet built: a dividends-received tally, a speculative-watch screen, per-name budgets for Fintech/Digital Platform, a manual price-entry box on the dashboard, and a poster version without profit-and-loss for showing Dad.', '2026-08-04'),
]

if __name__ == '__main__':
    out = sys.argv[1] if len(sys.argv) > 1 else os.path.join(REPO, '..', 'Pillar4_Stock_Notes.xlsx')
    out = os.path.abspath(out)
    path, n, nt, ns = build(out)
    print(f'{path} written: {n} comments across {nt} tickers, {ns} dashboard notes')
